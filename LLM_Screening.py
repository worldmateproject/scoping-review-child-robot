# (worldmate) C:\Users\m521633\OneDrive - Abertay University\Systematic Review\systematic_filter_result_based_code_And_ChatApi_code\ChatApi_code>python LLM_Screening_Only.py --model gpt-4.1-mini
"""
LLM_Screening.py
---------------------
Screens 1,899+ PDFs and outputs a 2-column related/not-related decision with justification,
targeting  eligible papers based on strict inclusion criteria.

Key changes from your previous extractor:
- Replaces structured data extraction with a binary screening stage.
- Avoids classification titles; only returns "Related" and "Justification".
- Enforces ALL eligibility criteria:
    1) Peer-reviewed journal or conference paper.
    2) Empirical methodology.
    3) Participants: typically developing children.
    4) Physical robot hardware used in the experiment.
    5) Focus on developmentally relevant learning outcomes.

Output Excel: Screening_Results.xlsx with columns:
[Model Used, paper_id, Related, Justification]
"""

import os
import json
import logging
import time
import argparse
from typing import List, Dict
import PyPDF2
import openai
import pandas as pd
from tenacity import retry, stop_after_attempt, wait_exponential
from dotenv import load_dotenv
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import re

# -------------------------------
# Logging
# -------------------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.FileHandler('screening.log'), logging.StreamHandler()]
)

# -------------------------------
# Configuration
# -------------------------------
MAX_TEXT_LENGTH = 150000     # ~120k characters (~30k tokens); keep consistent with your environment
API_MAX_RETRIES = 1
API_REQUEST_DELAY = 1        # Seconds between requests

# -------------------------------
# OpenAI Client
# -------------------------------
load_dotenv('API.env')
client = openai.OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
if not client.api_key:
    raise ValueError("API key not found in API.env file!")

# -------------------------------
# Helpers: Year & Age extraction
# -------------------------------
def extract_year(text: str) -> str:
    matches = re.findall(r"\b(?:19|20)\d{2}\b", text)
    if matches:
        # Choose the most frequent year mention (common in title pages/headers)
        return max(set(matches), key=matches.count)
    return "N/A"

def extract_age_from_text(text: str) -> str:
    """
    Regex-based extraction of age mentions from arbitrary paper text.
    Returns a semicolon-separated string of unique age patterns or "N/A".
    """
    age_patterns = [
        re.compile(
            r"""\b(?:aged|ages?)\s+(\d{1,2})
                 (?:\s*(?:-|–|—|to)\s*)(\d{1,2})\b
                 |
                 \bM\s*=\s*(\d{1,2}(?:\.\d+)?)""",
            re.IGNORECASE | re.VERBOSE
        ),
        re.compile(r"\bmean age\s*(?:is\s*)?(\d{1,2}(?:\.\d+)?)\b", re.IGNORECASE),
        re.compile(r"\bbetween\s+(\d{1,2})\s+(?:and|to)\s+(\d{1,2})\s+years?\b", re.IGNORECASE),
        re.compile(r"\b(\d{1,2})\s*years?\b", re.IGNORECASE),
    ]

    found_ages = set()
    for pattern in age_patterns:
        for match in pattern.findall(text):
            if isinstance(match, tuple):
                start_age = match[0] if len(match) > 0 else None
                end_age = match[1] if len(match) > 1 else None
                mean_age = match[2] if len(match) > 2 else None
                if start_age and end_age:
                    found_ages.add(f"{start_age}–{end_age} years")
                elif start_age:
                    found_ages.add(f"{start_age} years")
                elif mean_age:
                    found_ages.add(f"~{mean_age} years (mean)")
            else:
                found_ages.add(f"{match} years")

    return "; ".join(sorted(found_ages)) if found_ages else "N/A"

# -------------------------------
# Screening Prompt (System)
# -------------------------------
SCREENING_SYSTEM_PROMPT = (
  "You evaluate research papers for inclusion in a child–robot interaction (CRI) corpus.\n"
  "Return ONLY a flat JSON object with EXACTLY these 2 keys:\n"
  "  - \"Related\": one of [\"Yes\",\"No\"].\n"
  "  - \"Justification\": a single sentence (no line breaks) that cites decisive evidence from the text for why it is related or not related.\n"
  "\n"
  "Eligibility criteria (ALL must be satisfied for \"Yes\"):\n"
  "1) Peer-reviewed journal or conference paper.\n"
  "2) Empirical methodology (e.g., experiment, field study, RCT, case study with data); not a review, position paper, design fiction, conceptual, dataset-only, or purely technical without human participants.\n"
  "3) Participants are typically developing children (do NOT include clinical/special-needs-only samples unless a typically developing child cohort is clearly analyzed separately).\n"
  "4) A physical robot is used as hardware in the experiment (not only virtual agents/avatars/apps/videos; telepresence counts only if a physical robot body mediates the interaction).\n"
  "5) Primary focus is developmentally relevant learning outcomes (e.g., language/literacy, mathematics, science/CT, meta-learning/self-regulation, social–emotional development, collaboration/prosociality, health knowledge/habits, civic/green habits). Pure usability/likeability-only without a learning/development outcome is NOT eligible.\n"
  "\n"
  "Disqualify if ANY of the following applies:\n"
  "- Survey/review/editorial/position/vision; or purely technical (perception/control) without a human-child study.\n"
  "- Participants are adults, university students, or only clinical/special-needs without a distinct typically developing child cohort.\n"
  "- Only virtual agents or screens; no robot hardware used in the child study.\n"
  "- Outcomes focus only on system performance/usability without learning/development relevance.\n"
  "\n"
  "Output rules:\n"
  "A) Use ONLY double quotes; no markdown; no newlines inside values; no extra keys.\n"
  "B) If evidence is ambiguous, answer \"No\" and explain briefly.\n"
  "C) In \"Justification\", cite concrete cues (e.g., sample, ages, robot name, setting, outcome) that drove your decision.\n"
)

def create_screen_prompt(text: str) -> str:
    """
    Build a concise paper-level user prompt. No taxonomy titles, no extraction fields.
    """
    return f'PAPER TEXT:\n{text[:MAX_TEXT_LENGTH]}'

# -------------------------------
# PDF Text Extraction
# -------------------------------
def extract_text_from_pdf(file_path: str) -> str:
    """
    Extract text from all pages of the PDF, then truncate to MAX_TEXT_LENGTH.
    Sanitizes invalid surrogate characters.
    """
    try:
        with open(file_path, "rb") as f:
            pdf = PyPDF2.PdfReader(f)
            all_pages_text = []
            for page_num, page in enumerate(pdf.pages):
                try:
                    page_text = page.extract_text()
                    if page_text:
                        sanitized_text = page_text.encode('utf-16', 'surrogatepass').decode('utf-16', 'ignore')
                        all_pages_text.append(sanitized_text)
                except Exception as page_err:
                    logging.warning(f"Error extracting text from page {page_num} of {file_path}: {page_err}")
            combined_text = "\n".join(all_pages_text)
            return combined_text[:MAX_TEXT_LENGTH]
    except Exception as e:
        logging.error(f"PDF Error in {file_path}: {str(e)}")
        return ""

def load_papers(pdf_folder: str) -> List[Dict]:
    """Load and process PDF files."""
    pdfs = [f for f in os.listdir(pdf_folder) if f.lower().endswith(".pdf")]
    papers = []
    for f in pdfs:
        paper_id = os.path.splitext(f)[0]
        text = extract_text_from_pdf(os.path.join(pdf_folder, f))
        papers.append({"paper_id": paper_id, "text": text})
    return papers

# -------------------------------
# OpenAI Call: Screening Decision
# -------------------------------
@retry(stop=stop_after_attempt(API_MAX_RETRIES), wait=wait_exponential(multiplier=1, min=4, max=10))
def get_screening_decision(prompt: str, model: str) -> Dict:
    try:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": SCREENING_SYSTEM_PROMPT},
                {"role": "user", "content": prompt}
            ],
            temperature=0.0,
            top_p=1.0,
            response_format={"type": "json_object"}
        )
        content = response.choices[0].message.content.strip()
        data = json.loads(content)
        return {
            "Related": "Yes" if data.get("Related", "No") == "Yes" else "No",
            "Justification": data.get("Justification", "No decisive evidence found")
        }
    except Exception as e:
        logging.error(f"Screening API Error: {str(e)}")
        raise

# -------------------------------
# Processing Loop
# -------------------------------
def process_papers(papers: List[Dict], model: str) -> List[Dict]:
    """Run screening on each paper and collect results."""
    results = []
    for idx, paper in enumerate(papers):
        paper_id = paper["paper_id"]
        logging.info(f"Processing {idx+1}/{len(papers)}: {paper_id}")
        try:
            text = paper['text'] or ""
            # Optional boost: prepend regex signals to help the model
            extracted_year = extract_year(text)
            extracted_age = extract_age_from_text(text)
            if extracted_year != "N/A" or extracted_age != "N/A":
                boost = f"NOTE: year={extracted_year}; ages={extracted_age}\n\n"
            else:
                boost = ""

            prompt = create_screen_prompt(boost + text)
            decision = get_screening_decision(prompt, model)

            row = {
                "paper_id": paper_id,
                "Related": decision["Related"],
                "Justification": decision["Justification"]
            }
            results.append(row)
            time.sleep(API_REQUEST_DELAY)
        except Exception as e:
            logging.error(f"Failed {paper_id}: {str(e)}")
            results.append({
                "paper_id": paper_id,
                "Related": "No",
                "Justification": f"Error during screening: {str(e)}"
            })
    return results

# -------------------------------
# Excel Utilities
# -------------------------------
def sanitize_excel_string(value):
    """Remove illegal Excel characters from string values."""
    if isinstance(value, str):
        return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", value)
    return value

def sanitize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Sanitize all string cells."""
    return df.apply(lambda col: col.map(sanitize_excel_string) if col.dtype == "object" else col)

def save_to_excel(results: List[Dict], filename: str = "Screening_Results.xlsx") -> None:
    try:
        cols = ["Model Used", "paper_id", "Related", "Justification"]
        df = pd.DataFrame(results).replace(["", None], "N/A")
        if "Model Used" not in df.columns:
            df["Model Used"] = "N/A"
        df = df.reindex(columns=cols).fillna("N/A")
        df = sanitize_dataframe(df)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Screening')
            ws = writer.sheets['Screening']

            wrap = Alignment(wrap_text=True, vertical='top', horizontal='left')
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = wrap

            for idx, col in enumerate(df.columns, 1):
                max_len = min(max(df[col].astype(str).map(len).max(), len(str(col))) + 2, 60)
                ws.column_dimensions[get_column_letter(idx)].width = max_len

            for cell in ws[1]:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
                cell.font = Font(bold=True)

            ws.freeze_panes = 'A2'
    except Exception as e:
        logging.error(f"Error writing Excel file: {e}")
        raise

# -------------------------------
# Main
# -------------------------------
def main():
    parser = argparse.ArgumentParser(description="CRI Paper Screening (Related/Not Related)")
    parser.add_argument("--model", type=str, default="gpt-4.1-mini", help="OpenAI model to use for screening")
    parser.add_argument("--pdf_folder", type=str, default="paperpdf", help="Folder containing PDF files")
    parser.add_argument("--out", type=str, default="Screening_Results.xlsx", help="Output Excel filename")
    args = parser.parse_args()

    if not os.path.exists(args.pdf_folder):
        logging.error(f"PDF folder not found: {args.pdf_folder}")
        return

    papers = load_papers(args.pdf_folder)
    if not papers:
        logging.error("No valid PDFs found")
        return

    logging.info(f"Starting screening with {args.model} on {len(papers)} papers")
    results = process_papers(papers, args.model)

    for r in results:
        r["Model Used"] = args.model

    save_to_excel(results, filename=args.out)
    logging.info(f"Processing complete. Data saved to {args.out}")

if __name__ == "__main__":
    main()
